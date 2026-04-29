import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk, simpledialog
import pdfplumber
import pandas as pd
import re
import os
import threading
import json
import ctypes 
from datetime import datetime
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl import load_workbook

# CONFIGURACIÓN VISUAL (HIGH DPI)
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass 

# GESTIÓN DE ARCHIVOS
ARCHIVO_MEMORIA = "diccionario_productos.json"
ARCHIVO_CATEGORIAS = "config_categorias.json"
ARCHIVO_LOG = "registro_actividad.log"

CATEGORIAS_DEFAULT = [
    "EQUIPO",
    "ACCESORIO",
    "ACCESORIO SERIE",
    "SIMCA",
    "SIMCA SERIE",
    "PROM",
    "PROM SERIE",
    "FICHA",
    "FICHA SERIE",
    "PUBL",
    "PUBL SERIE",
    "JUGUE",
    "IOT",
    "IOT SERIE",
    "TAE"
]

def cargar_json(archivo, default_data):
    if os.path.exists(archivo):
        try:
            with open(archivo, 'r', encoding='utf-8') as f:
                return json.load(f)
        except: return default_data
    return default_data

def guardar_json(archivo, data):
    try:
        with open(archivo, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except: pass

def obtener_categorias():
    return cargar_json(ARCHIVO_CATEGORIAS, CATEGORIAS_DEFAULT)

# GESTOR DE CATEGORÍAS
class GestorCategorias:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Administrar Categorías")
        self.top.geometry("450x500")
        
        frame_input = tk.Frame(self.top)
        frame_input.pack(pady=10, fill=tk.X, padx=10)
        tk.Label(frame_input, text="Nueva:").pack(side=tk.LEFT)
        self.entry_cat = tk.Entry(frame_input, width=25)
        self.entry_cat.pack(side=tk.LEFT, padx=5)
        tk.Button(frame_input, text="Agregar", command=self.agregar, bg="#0070C0", fg="white").pack(side=tk.LEFT)

        frame_list = tk.Frame(self.top)
        frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scrollbar = tk.Scrollbar(frame_list)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox = tk.Listbox(frame_list, yscrollcommand=scrollbar.set, selectmode=tk.SINGLE, height=15)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)

        frame_btn = tk.Frame(self.top)
        frame_btn.pack(pady=10)
        tk.Button(frame_btn, text="Renombrar", command=self.renombrar, bg="#FFC000", fg="black").pack(side=tk.LEFT, padx=5)
        tk.Button(frame_btn, text="Eliminar", command=self.eliminar, bg="#C00000", fg="white").pack(side=tk.LEFT, padx=5)
        tk.Button(frame_btn, text="Restaurar", command=self.restaurar, bg="gray", fg="white").pack(side=tk.LEFT, padx=5)

        self.cargar_lista()

    def cargar_lista(self):
        self.listbox.delete(0, tk.END)
        cats = obtener_categorias()
        for c in sorted(cats): self.listbox.insert(tk.END, c)

    def agregar(self):
        nueva = self.entry_cat.get().strip()
        if nueva:
            cats = obtener_categorias()
            if nueva not in cats:
                cats.append(nueva)
                guardar_json(ARCHIVO_CATEGORIAS, cats)
                self.cargar_lista()
                self.entry_cat.delete(0, tk.END)

    def renombrar(self):
        sel = self.listbox.curselection()
        if not sel: return
        viejo = self.listbox.get(sel[0])
        nuevo = simpledialog.askstring("Renombrar", f"Renombrar '{viejo}' a:", parent=self.top, initialvalue=viejo)
        if nuevo and nuevo.strip():
            nuevo = nuevo.strip()
            cats = obtener_categorias()
            cats = [nuevo if x == viejo else x for x in cats]
            guardar_json(ARCHIVO_CATEGORIAS, cats)
            
            memoria = cargar_json(ARCHIVO_MEMORIA, {})
            cnt = 0
            for k, v in memoria.items():
                if v == viejo: 
                    memoria[k] = nuevo
                    cnt += 1
            if cnt > 0: guardar_json(ARCHIVO_MEMORIA, memoria)
            messagebox.showinfo("Listo", f"Categoría renombrada. {cnt} productos actualizados.")
            self.cargar_lista()

    def eliminar(self):
        sel = self.listbox.curselection()
        if not sel: return
        val = self.listbox.get(sel[0])
        
        if messagebox.askyesno("Confirmar", f"¿Eliminar categoría '{val}'?\n\nEsto borrará también los productos asociados en la memoria."):
            cats = obtener_categorias()
            if val in cats:
                cats.remove(val)
                guardar_json(ARCHIVO_CATEGORIAS, cats)

            memoria = cargar_json(ARCHIVO_MEMORIA, {})
            productos_a_borrar = [k for k, v in memoria.items() if v == val]
            for k in productos_a_borrar: del memoria[k]
            
            if productos_a_borrar:
                guardar_json(ARCHIVO_MEMORIA, memoria)
                
            self.cargar_lista()
            messagebox.showinfo("Eliminado", f"Categoría eliminada y memoria limpiada.")

    def restaurar(self):
        if messagebox.askyesno("Confirmar", "¿Restaurar defaults?"):
            guardar_json(ARCHIVO_CATEGORIAS, CATEGORIAS_DEFAULT)
            self.cargar_lista()

# EDITOR MEMORIA
class EditorMemoria:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Gestor de Productos")
        self.top.geometry("800x500")
        
        frame_tools = tk.Frame(self.top)
        frame_tools.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        tk.Label(frame_tools, text="Buscar:").pack(side=tk.LEFT)
        self.entry_search = tk.Entry(frame_tools, width=25)
        self.entry_search.pack(side=tk.LEFT, padx=5)
        self.entry_search.bind("<KeyRelease>", self.actualizar_vista)
        
        tk.Label(frame_tools, text="Ordenar por:").pack(side=tk.LEFT, padx=(15, 0))
        self.combo_sort = ttk.Combobox(frame_tools, values=[
            "Más Recientes Primero", 
            "Más Antiguos Primero", 
            "A-Z (Alfabético)", 
            "Z-A (Inverso)"
        ], state="readonly", width=20)
        self.combo_sort.pack(side=tk.LEFT, padx=5)
        self.combo_sort.current(0)
        self.combo_sort.bind("<<ComboboxSelected>>", self.actualizar_vista)

        frame_btn = tk.Frame(self.top)
        frame_btn.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        tk.Button(frame_btn, text="Editar Seleccionado", command=self.editar_item, bg="#0070C0", fg="white").pack(side=tk.LEFT, padx=5)
        tk.Button(frame_btn, text="Olvidar Seleccionado", command=self.eliminar_item, bg="#C00000", fg="white").pack(side=tk.LEFT, padx=5)
        self.lbl_count = tk.Label(frame_btn, text="0 productos", fg="gray")
        self.lbl_count.pack(side=tk.RIGHT)

        cols = ("Producto", "Categoría")
        self.tree = ttk.Treeview(self.top, columns=cols, show='headings')
        self.tree.heading("Producto", text="Descripción")
        self.tree.heading("Categoría", text="Categoría")
        self.tree.column("Producto", width=500)
        
        scrollbar = tk.Scrollbar(self.top, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
        
        self.memoria_completa = {}
        self.cargar_datos_iniciales()

    def cargar_datos_iniciales(self):
        self.memoria_completa = cargar_json(ARCHIVO_MEMORIA, {})
        self.actualizar_vista()

    def actualizar_vista(self, event=None):
        for i in self.tree.get_children(): self.tree.delete(i)
        
        termino = self.entry_search.get().lower()
        modo_sort = self.combo_sort.get()
        
        items_filtrados = []
        raw_items = list(self.memoria_completa.items())
        
        for prod, cat in raw_items:
            if termino in prod.lower() or termino in cat.lower():
                items_filtrados.append((prod, cat))
        
        if modo_sort == "A-Z (Alfabético)":
            items_filtrados.sort(key=lambda x: x[0].lower())
        elif modo_sort == "Z-A (Inverso)":
            items_filtrados.sort(key=lambda x: x[0].lower(), reverse=True)
        elif modo_sort == "Más Recientes Primero":
            items_filtrados.reverse()
        elif modo_sort == "Más Antiguos Primero":
            pass

        for prod, cat in items_filtrados:
            self.tree.insert("", "end", values=(prod, cat))
            
        self.lbl_count.config(text=f"{len(items_filtrados)} productos encontrados")

    def eliminar_item(self):
        sel = self.tree.selection()
        if sel:
            item = self.tree.item(sel[0])
            prod = item['values'][0]
            if messagebox.askyesno("Confirmar", f"¿Olvidar '{prod}'?"):
                if prod in self.memoria_completa:
                    del self.memoria_completa[prod]
                    guardar_json(ARCHIVO_MEMORIA, self.memoria_completa)
                    self.actualizar_vista()

    def editar_item(self):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        prod, cat = item['values'][0], item['values'][1]
        
        win = tk.Toplevel(self.top)
        win.title("Editar")
        tk.Label(win, text=prod, fg="blue").pack(pady=5)
        combo = ttk.Combobox(win, values=obtener_categorias())
        combo.set(cat)
        combo.pack(pady=5)
        def save():
            new_cat = combo.get()
            if new_cat:
                self.memoria_completa[prod] = new_cat
                guardar_json(ARCHIVO_MEMORIA, self.memoria_completa)
                self.actualizar_vista()
                win.destroy()
        tk.Button(win, text="Guardar", command=save).pack(pady=10)

# VISOR DE HISTORIAL
class VisorHistorial:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Historial por Días")
        self.top.geometry("700x500")

        paned = tk.PanedWindow(self.top, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        frame_left = tk.Frame(paned)
        paned.add(frame_left, width=200)
        tk.Label(frame_left, text="Fechas Disponibles", font=("Segoe UI", 9, "bold")).pack(pady=5)
        self.tree_dates = ttk.Treeview(frame_left, columns=("Fecha"), show="headings")
        self.tree_dates.heading("Fecha", text="Día")
        self.tree_dates.pack(fill=tk.BOTH, expand=True)
        self.tree_dates.bind("<<TreeviewSelect>>", self.mostrar_detalle)

        frame_right = tk.Frame(paned)
        paned.add(frame_right)
        tk.Label(frame_right, text="Registro de Actividad", font=("Segoe UI", 9, "bold")).pack(pady=5)
        self.txt_detail = scrolledtext.ScrolledText(frame_right, font=("Consolas", 9), state='disabled')
        self.txt_detail.pack(fill=tk.BOTH, expand=True)

        self.logs_por_fecha = {}
        self.cargar_logs()

    def cargar_logs(self):
        if not os.path.exists(ARCHIVO_LOG): return
        self.logs_por_fecha = {}
        try:
            with open(ARCHIVO_LOG, "r", encoding="utf-8") as f:
                lines = f.readlines()
            fecha_actual = "Sin Fecha"
            for line in lines:
                match = re.search(r'\[(\d{4}-\d{2}-\d{2})', line)
                if match: fecha_actual = match.group(1)
                if fecha_actual not in self.logs_por_fecha: self.logs_por_fecha[fecha_actual] = []
                self.logs_por_fecha[fecha_actual].append(line)
            
            fechas_ordenadas = sorted(self.logs_por_fecha.keys(), reverse=True)
            for fecha in fechas_ordenadas:
                if fecha != "Sin Fecha": self.tree_dates.insert("", "end", values=(fecha,))
            if "Sin Fecha" in self.logs_por_fecha:
                 self.tree_dates.insert("", "end", values=("Sin Fecha",))
        except Exception as e: messagebox.showerror("Error", f"No se pudo leer el log: {e}")

    def mostrar_detalle(self, event):
        sel = self.tree_dates.selection()
        if not sel: return
        fecha = self.tree_dates.item(sel[0])['values'][0]
        contenido = self.logs_por_fecha.get(str(fecha), [])
        self.txt_detail.config(state='normal')
        self.txt_detail.delete(1.0, tk.END)
        for line in contenido: self.txt_detail.insert(tk.END, line)
        self.txt_detail.config(state='disabled')

# LÓGICA DE NEGOCIO
def limpiar_documento(doc_str):
    if doc_str and doc_str.startswith("80") and len(doc_str) > 8: return doc_str[2:]
    return doc_str

def buscar_uuid_definitivo(texto_completo):
    if not texto_completo: return ""
    texto_plano = texto_completo.replace(" ", "").replace("\n", "").replace("\r", "")
    patron = r'(?i)[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'
    coincidencias = re.findall(patron, texto_plano)
    if coincidencias:
        return coincidencias[0].lower()
    return ""

def aplicar_estilo_visual(ruta_excel):
    try:
        wb = load_workbook(ruta_excel)
        ws = wb.active
        
        side_vacio = Side(border_style=None)
        no_border = Border(left=side_vacio, right=side_vacio, top=side_vacio, bottom=side_vacio)
        
        font_header = Font(name='Calibri', size=11, color="C65911", bold=False)
        font_blue = Font(name='Calibri', size=11, color="0070C0")
        font_normal = Font(name='Calibri', size=11, color="000000")
        
        no_fill = PatternFill(fill_type=None)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = no_border
                cell.fill = no_fill
                cell.number_format = '@' 
                
                if cell.row == 1:
                    cell.font = font_header
                else:
                    if cell.column in [1, 2, 3, 7, 10]:
                        cell.font = font_blue
                    else:
                        cell.font = font_normal

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['D'].width = 40 
        ws.column_dimensions['G'].width = 25 
        ws.column_dimensions['J'].width = 40 
        
        wb.save(ruta_excel)
    except Exception as e:
        print(f"Error al formatear: {e}")

# APP PRINCIPAL
class ExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Extractor Facturas Telcel PRO")
        self.root.geometry("700x550")
        
        self.user_choice = None
        self.event_wait = threading.Event()

        # Menú
        menubar = tk.Menu(root)
        root.config(menu=menubar)
        tools = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Herramientas", menu=tools)
        tools.add_command(label="Administrar Categorías", command=lambda: GestorCategorias(self.root))
        tools.add_command(label="Editar Memoria", command=lambda: EditorMemoria(self.root))
        tools.add_separator()
        tools.add_command(label="Salir", command=root.quit)

        # UI
        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="Convertidor PDF a Excel", font=("Segoe UI", 18, "bold")).pack(pady=(0, 10))
        
        btn_frame = tk.Frame(frame)
        btn_frame.pack(pady=10)
        
        self.btn_select = tk.Button(btn_frame, text="Seleccionar PDFs", font=("Segoe UI", 11), 
                                    bg="#0070C0", fg="white", width=25, height=2, command=self.iniciar_hilo)
        self.btn_select.pack(side=tk.LEFT, padx=5)

        # Frame Logs
        frame_log = tk.Frame(frame)
        frame_log.pack(fill=tk.BOTH, expand=True, pady=10)
        
        header_log = tk.Frame(frame_log)
        header_log.pack(fill=tk.X)
        tk.Label(header_log, text="Historial de Actividad:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        frame_hist_btns = tk.Frame(header_log)
        frame_hist_btns.pack(side=tk.RIGHT)
        tk.Button(frame_hist_btns, text="📅 Ver Historial por Días", font=("Segoe UI", 8), 
                  command=lambda: VisorHistorial(self.root), bg="#FFC000").pack(side=tk.LEFT, padx=5)
        tk.Button(frame_hist_btns, text="🗑 Borrar Todo", font=("Segoe UI", 8), 
                  command=self.borrar_historial, bg="#EEE").pack(side=tk.LEFT)

        self.txt_log = scrolledtext.ScrolledText(frame_log, height=12, font=("Consolas", 9), state='disabled')
        self.txt_log.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(frame, text="Versión 11.13 - High Capacity (50k Range)", fg="gray").pack(pady=5)
        
        self.log("--- Aplicación Iniciada ---")

    def log(self, mensaje):
        ahora = datetime.now().strftime("[%Y-%m-%d %H:%M:%S] ")
        texto_completo = f"{ahora}{mensaje}"
        self.txt_log.config(state='normal')
        self.txt_log.insert(tk.END, texto_completo + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state='disabled')
        try:
            with open(ARCHIVO_LOG, "a", encoding="utf-8") as f:
                f.write(texto_completo + "\n")
        except: pass

    def borrar_historial(self):
        if messagebox.askyesno("Confirmar", "¿Borrar permanentemente TODO el historial?"):
            self.txt_log.config(state='normal')
            self.txt_log.delete(1.0, tk.END)
            self.txt_log.config(state='disabled')
            try:
                open(ARCHIVO_LOG, 'w').close()
            except: pass

    def iniciar_hilo(self):
        rutas_pdf = filedialog.askopenfilenames(title="Seleccionar Facturas (Una o Varias)", filetypes=[("PDF", "*.pdf")])
        if not rutas_pdf: return
        
        self.btn_select.config(state="disabled", text="Procesando...")
        self.log(f"--- Inicio Sesión: {len(rutas_pdf)} archivos seleccionados ---")
        
        t = threading.Thread(target=self.procesar_logica, args=(rutas_pdf,))
        t.start()

    def solicitar_clasificacion(self, descripcion, codigo):
        self.user_choice = None
        self.event_wait.clear()
        self.root.after(0, lambda: self._abrir_popup(descripcion, codigo))
        self.event_wait.wait()
        return self.user_choice

    def _abrir_popup(self, descripcion, codigo):
        top = tk.Toplevel(self.root)
        top.title("Clasificar Producto")
        top.geometry("450x300")
        top.attributes('-topmost', True)
        
        tk.Label(top, text="Producto Nuevo (o Categoría Eliminada):", font=("Segoe UI", 10, "bold")).pack(pady=10)
        tk.Label(top, text=f"{descripcion}", fg="blue", wraplength=400).pack()
        
        cats = obtener_categorias()
        combo = ttk.Combobox(top, values=cats, width=35)
        combo.pack(pady=10)
        if cats: combo.current(0)
        
        def guardar():
            val = combo.get().strip()
            if val:
                curr = obtener_categorias()
                if val not in curr:
                    if messagebox.askyesno("Nueva", f"¿Agregar '{val}' a la lista de categorías?"):
                        curr.append(val)
                        guardar_json(ARCHIVO_CATEGORIAS, curr)
                self.user_choice = val
                self.event_wait.set()
                top.destroy()
        
        def cerrar():
            self.user_choice = obtener_categorias()[0] if obtener_categorias() else "ACCESORIO"
            self.event_wait.set()
            top.destroy()

    def procesar_logica(self, pdf_paths):
        try:
            memoria = cargar_json(ARCHIVO_MEMORIA, {})
            categorias_validas = obtener_categorias()
            
            exitos = 0
            carpeta_destino = None

            if len(pdf_paths) > 1:
                self.log("Solicitando carpeta de destino...")
                carpeta_destino = filedialog.askdirectory(title=f"Selecciona carpeta para guardar {len(pdf_paths)} archivos")
                
                if not carpeta_destino:
                    self.log("Operación cancelada: No se seleccionó carpeta.")
                    messagebox.showwarning("Cancelado", "No se seleccionó una carpeta de destino.\nSe canceló el proceso.")
                    return 
            
            for index, pdf_path in enumerate(pdf_paths):
                nombre_archivo = os.path.basename(pdf_path)
                self.log(f"[{index+1}/{len(pdf_paths)}] Analizando: {nombre_archivo}...")
                
                datos_archivo = []
                cambios = False

                try:
                    with pdfplumber.open(pdf_path) as pdf:
                        lines = []
                        raw_text_full = ""
                        for p in pdf.pages:
                            l = p.extract_text(layout=True)
                            r = p.extract_text() 
                            if l: lines.extend(l.split('\n'))
                            if r: raw_text_full += r + "\n"

                        uuid = buscar_uuid_definitivo(raw_text_full)
                        
                        doc = ""
                        for i in range(min(60, len(lines))):
                            if "DOCUMENTO" in lines[i]:
                                if i+1 < len(lines) and len(lines[i+1].split()[0]) >= 8:
                                    doc = limpiar_documento(lines[i+1].split()[0])
                                    break
                        if not doc:
                             m = re.search(r'(\d+)', nombre_archivo)
                             if m: doc = limpiar_documento(m.group(1))

                        patron = r'(\d{7,8})\s+(.+?)\s+([\d,]+)\s+PZA\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)'
                        
                        for i, line in enumerate(lines):
                            m_prod = re.search(patron, line)
                            if m_prod:
                                cod = m_prod.group(1)
                                desc = re.sub(r'\s+', ' ', m_prod.group(2).replace('#', ' ')).strip()
                                try:
                                    cant = int(m_prod.group(3).replace(',', ''))
                                    precio = float(m_prod.group(4).replace(',', ''))
                                    dcto = float(m_prod.group(5).replace(',', ''))
                                    udcto = (dcto / cant) if dcto > precio else dcto
                                    final = round(max(0, precio - udcto) * 1.16, 4)
                                except: continue

                                series = []
                                cat_sat = None
                                offset = 1
                                while offset < 15000: 
                                    if i+offset >= len(lines): break
                                    nline = lines[i+offset]
                                    if re.search(patron, nline): break
                                    if any(x in nline.upper() for x in ["SUBTOTAL", "TOTAL", "SELLO DIGITAL"]): break
                                    
                                    upper = nline.upper()
                                    if "CLAVE PROD O SERV" in upper:
                                        if "TELEFONOS" in upper: cat_sat = "EQUIPO"
                                        elif "PROMOCIONAL" in upper: cat_sat = "PROM"
                                        elif "ACCESORIOS" in upper: 
                                            cat_sat = "SIMCA SERIE" if "SIM" in desc.upper() else "ACCESORIO"
                                    
                                    mr = re.finditer(r'(?:(\d{3,8})\s+)?(\d{10,20})\s*-\s*(\d{10,20})', nline)
                                    found_r = False
                                    for m in mr:
                                        found_r = True
                                        try:
                                            ini, fin = int(m.group(2)), int(m.group(3))
                                            tot = fin - ini + 1
                                            if 0 < tot <= 50000:
                                                for k in range(tot): series.append(str(ini + k))
                                        except: pass
                                    
                                    if not found_r:
                                        ms = re.findall(r'(?<!\d)(\d{15,20})(?!\d)', nline)
                                        for s in ms: 
                                            if not s.startswith("00001000"): series.append(s)
                                    offset += 1

                                tipo = "DESC"
                                if desc in memoria: 
                                    tipo_memoria = memoria[desc]
                                    if tipo_memoria in categorias_validas:
                                        tipo = tipo_memoria
                                    else:
                                        del memoria[desc]
                                        cambios = True
                                        tipo = "DESC"

                                if tipo == "DESC" and cat_sat:
                                    if cat_sat in categorias_validas:
                                        tipo = cat_sat
                                        if tipo == "PROM" and series: 
                                            tipo = "PROM SERIE" if "PROM SERIE" in categorias_validas else tipo
                                        if tipo == "SIMCA SERIE" and not series: 
                                            tipo = "SIMCA" if "SIMCA" in categorias_validas else tipo
                                        if tipo == "ACCESORIO" and series: 
                                            tipo = "ACCESORIO SERIE" if "ACCESORIO SERIE" in categorias_validas else tipo
                                    else:
                                        tipo = "DESC"

                                if tipo == "DESC":
                                    if series: 
                                        if "SIM" in desc.upper(): cand = "SIMCA SERIE"
                                        else: cand = "EQUIPO"
                                    else:
                                        if "FICHA" in desc.upper(): cand = "FICHA"
                                        elif "PROM" in desc.upper(): cand = "PROM"
                                        elif "PREPAGO" in desc.upper(): cand = "PROM"
                                        else: cand = "ACCESORIO"
                                    
                                    if cand in categorias_validas:
                                        tipo = cand
                                    else:
                                        tipo = "DESC"
                                
                                if tipo == "DESC":
                                    tipo = self.solicitar_clasificacion(desc, cod)
                                    memoria[desc] = tipo
                                    cambios = True
                                
                                if desc not in memoria and tipo != "DESC":
                                    memoria[desc] = tipo
                                    cambios = True

                                cnt = 0
                                for s in series[:cant]:
                                    datos_archivo.append([tipo, doc, cod, desc, 1, "PZA", s, "NA", final, uuid, 0])
                                    cnt += 1
                                rest = cant - cnt
                                if rest > 0:
                                     datos_archivo.append([tipo, doc, cod, desc, rest, "PZA", "NA", "NA", final, uuid, 0])

                    if datos_archivo:
                        df = pd.DataFrame(datos_archivo, columns=["Tipo", "documento", "codigo Articulo", "descripcion articulo", "pzas articulo", "unidad medida", "serie1", "serie2", "precio unitario", "uuid SAT", "IDArticulo INTRA"])
                        
                        ruta_guardar = None
                        nombre_base = os.path.splitext(nombre_archivo)[0]
                        if nombre_base.startswith("80"): nombre_base = nombre_base[2:]

                        if len(pdf_paths) == 1:
                            ruta_guardar = filedialog.asksaveasfilename(
                                defaultextension=".xlsx",
                                filetypes=[("Excel", "*.xlsx")],
                                initialfile=f"{nombre_base}.xlsx",
                                title="Guardar Excel Como..."
                            )
                        else:
                            if carpeta_destino:
                                ruta_guardar = os.path.join(carpeta_destino, f"{nombre_base}.xlsx")
                        
                        if ruta_guardar:
                            df.to_excel(ruta_guardar, index=False)
                            aplicar_estilo_visual(ruta_guardar)
                            self.log(f"   -> Guardado: {os.path.basename(ruta_guardar)}")
                            exitos += 1
                        else:
                            self.log("   -> Guardado cancelado por usuario.")
                    else:
                        self.log("   -> [AVISO] No se encontraron datos en este archivo.")

                except Exception as e:
                    self.log(f"   -> [ERROR] Falló {nombre_archivo}: {str(e)}")

            if cambios: 
                guardar_json(ARCHIVO_MEMORIA, memoria)
                self.log("Memoria actualizada globalmente.")

            messagebox.showinfo("Proceso Terminado", f"Se procesaron {len(pdf_paths)} archivos.\nSe generaron {exitos} Excels exitosamente.")

        except Exception as e:
            self.log(f"[ERROR CRÍTICO] {str(e)}")
            messagebox.showerror("Error", str(e))
        finally:
            self.btn_select.config(state="normal", text="Seleccionar PDFs")

if __name__ == "__main__":
    root = tk.Tk()
    try:
        root.iconbitmap(default='icono.ico')
    except:
        pass
    app = ExtractorApp(root)
    root.mainloop()